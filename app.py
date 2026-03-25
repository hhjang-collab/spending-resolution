import streamlit as st
import pandas as pd
import base64
import os
import io
from datetime import datetime
import openpyxl

# 1. 페이지 기본 설정 (공통 필수 규칙 2)
st.set_page_config(page_title="지출결의서 작성 앱", layout="centered")

# 2. UI 최적화 CSS (공통 필수 규칙 6)
st.markdown("""
    <style>
        [data-testid="InputInstructions"] {display: none !important;}
    </style>
""", unsafe_allow_html=True)

# 3. 보안 (비밀번호) 로그인 로직 (공통 필수 규칙 3)
# [나중에 직접 채워 넣어야 하는 부분] Streamlit Cloud 대시보드의 Secrets에 APP_PASSWORD = "비밀번호" 를 입력하세요.
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.title("🔒 로그인")
    st.markdown("사내 업무용 시스템입니다. 비밀번호를 입력해주세요.")
    pwd_input = st.text_input("비밀번호", type="password")
    
    if st.button("로그인"):
        correct_password = st.secrets.get("APP_PASSWORD", "1234") 
        if pwd_input == correct_password:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("비밀번호가 일치하지 않습니다.")
    st.stop()

# 4. 회사 로고 (우측 상단 고정) (공통 필수 규칙 4)
def get_base64_of_bin_file(bin_file):
    if os.path.exists(bin_file):
        with open(bin_file, 'rb') as f:
            data = f.read()
        return base64.b64encode(data).decode()
    return ""

# [나중에 직접 채워 넣어야 하는 부분] 깃허브 저장소에 "company_logo.png" 파일을 업로드해 두세요.
logo_base64 = get_base64_of_bin_file("company_logo.png")
if logo_base64:
    st.markdown(f"""
        <style>
            .fixed-logo {{
                position: fixed;
                top: 70px;
                right: 30px;
                z-index: 999;
                width: 120px;
            }}
            @media (max-width: 768px) {{
                .fixed-logo {{
                    top: 15px;
                    right: 15px;
                    width: 80px;
                }}
            }}
        </style>
        <img src="data:image/png;base64,{logo_base64}" class="fixed-logo">
    """, unsafe_allow_html=True)

# 5. 홈 버튼 (포털 복귀) 및 얇은 여백 구분선 (공통 필수 규칙 5)
with st.sidebar:
    st.markdown(
        '''
        <div style="margin-top: 5px;">
            <a href="https://ip2b-work-tools.streamlit.app/" target="_blank" style="text-decoration: none; color: #31333F; font-size: 15px; font-weight: 600;">
                🏠 홈으로
            </a>
        </div>
        <hr style="margin-top: 10px; margin-bottom: 15px; border: 0; border-top: 1px solid rgba(49, 51, 63, 0.2);">
        ''', 
        unsafe_allow_html=True
    )
    st.markdown("### 📝 지출결의 메뉴")
    st.button("신규 작성", use_container_width=True)
    st.button("작성 내역 조회", use_container_width=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### ⚙️ 관리자 메뉴")
    st.button("📊 월별 전체내역 엑셀 다운로드", use_container_width=True)

# 6. 여백이 얇은 구분선 함수화 (공통 필수 규칙 7)
def thin_divider():
    st.markdown('<hr style="margin-top: 15px; margin-bottom: 15px; border: 0; border-top: 1px solid rgba(49, 51, 63, 0.2);">', unsafe_allow_html=True)

# --- 상태 초기화 (Session State) ---
if "form_data" not in st.session_state:
    st.session_state["form_data"] = {
        "project": "선택", "department": "기술사업화팀", "author": "",
        "account": "선택", "purpose": "", "title": "대리", "date": datetime.today()
    }
if "expense_data" not in st.session_state:
    st.session_state["expense_data"] = pd.DataFrame(
        [{"지출일": datetime.today().date(), "적요": "식대", "지급처": "", "금액": 0, "결제구분": "법인카드", "첨부": "매출전표", "비고": ""}]
    )

# --- 메인 화면 로직 ---
st.title("📄 지출결의서 작성")
st.caption("작성된 데이터는 깃허브에 저장된 기존 엑셀 양식 서식을 그대로 유지한 채 출력됩니다.")

# --- 기존 작성본 업로드 (자동 완성 영역) ---
uploaded_file = st.file_uploader("📂 기존에 작성했던 엑셀 파일이 있다면 업로드하세요. (내용 자동 입력)", type=['xlsx'])

if uploaded_file is not None:
    try:
        df_up = pd.read_excel(uploaded_file, header=None)
        st.session_state["form_data"]["project"] = str(df_up.iloc[5, 1]) if pd.notna(df_up.iloc[5, 1]) else "선택"
        st.session_state["form_data"]["purpose"] = str(df_up.iloc[6, 1]) if pd.notna(df_up.iloc[6, 1]) else ""
        st.session_state["form_data"]["department"] = str(df_up.iloc[7, 1]) if pd.notna(df_up.iloc[7, 1]) else ""
        st.session_state["form_data"]["title"] = str(df_up.iloc[7, 3]) if pd.notna(df_up.iloc[7, 3]) else ""
        st.session_state["form_data"]["author"] = str(df_up.iloc[8, 1]) if pd.notna(df_up.iloc[8, 1]) else ""
        st.session_state["form_data"]["account"] = str(df_up.iloc[7, 5]) if pd.notna(df_up.iloc[7, 5]) else "선택"
        
        ex_list = []
        for i in range(11, len(df_up)):
            val = df_up.iloc[i, 0]
            if pd.isna(val): continue
            ex_list.append({
                "지출일": pd.to_datetime(val).date() if isinstance(val, str) else val,
                "적요": str(df_up.iloc[i, 1]) if pd.notna(df_up.iloc[i, 1]) else "식대",
                "지급처": str(df_up.iloc[i, 3]) if pd.notna(df_up.iloc[i, 3]) else "",
                "금액": int(df_up.iloc[i, 5]) if pd.notna(df_up.iloc[i, 5]) else 0,
                "결제구분": "법인카드",
                "첨부": "매출전표",
                "비고": str(df_up.iloc[i, 7]) if pd.notna(df_up.iloc[i, 7]) else ""
            })
        if ex_list: st.session_state["expense_data"] = pd.DataFrame(ex_list)
        st.toast("✅ 기존 데이터 불러오기 성공!", icon="✨")
    except Exception as e:
        st.error(f"파일 읽기 오류: {e}")

thin_divider()

# 첨부 데이터를 기반으로 한 선택지
PROJECT_LIST = ["선택", "전북 군산산단 AX마스터플랜 수립 연구", "상주시-글로컬", "KEITI-중소환경", "대한의협-정보화", "NIPA-SW인재", "호서대-지역청년", "대구TP-과학치안", "KNU-진단 2024"]
ACCOUNT_LIST = ["선택", "출장비", "회의비", "복리후생비"]
SUMMARY_LIST = ["식대", "다과비", "교통비", "교통비(KTX)", "교통비(카셰어링)", "주유비", "숙박비", "간식비"]
PAYMENT_METHODS = ["현금", "계좌이체", "법인카드", "개인카드", "복합결제", "사업비카드"]
ATTACHMENTS = ["매출전표", "세금계산서", "현금영수증"]

def get_idx(lst, item): return lst.index(item) if item in lst else 0

st.subheader("📝 기본 정보")
col1, col2 = st.columns(2)
with col1:
    project = st.selectbox("프로젝트", PROJECT_LIST, index=get_idx(PROJECT_LIST, st.session_state["form_data"]["project"]))
    department = st.text_input("부서", value=st.session_state["form_data"]["department"])
    author = st.text_input("작성자", value=st.session_state["form_data"]["author"])
    account = st.selectbox("계정과목", ACCOUNT_LIST, index=get_idx(ACCOUNT_LIST, st.session_state["form_data"]["account"]))
with col2:
    purpose = st.text_input("목적", value=st.session_state["form_data"]["purpose"])
    title = st.text_input("직위", value=st.session_state["form_data"]["title"])
    date = st.date_input("작성일", st.session_state["form_data"]["date"])
    
thin_divider()

st.subheader("💳 지출 내역 상세")
edited_df = st.data_editor(
    st.session_state["expense_data"],
    column_config={
        "지출일": st.column_config.DateColumn("지출일", required=True),
        "적요": st.column_config.SelectboxColumn("적요", options=SUMMARY_LIST, required=True),
        "지급처": st.column_config.TextColumn("지급처", required=True),
        "금액": st.column_config.NumberColumn("금액", min_value=0, step=1000, required=True, format="%d 원"),
        "결제구분": st.column_config.SelectboxColumn("결제구분", options=PAYMENT_METHODS, required=True),
        "첨부": st.column_config.SelectboxColumn("첨부", options=ATTACHMENTS, required=True),
        "비고": st.column_config.TextColumn("비고"),
    },
    num_rows="dynamic", use_container_width=True, hide_index=True
)

total_amount = edited_df["금액"].sum()
st.markdown(f"**총 지출 금액: <span style='color:#e74c3c;'>{total_amount:,} 원</span>**", unsafe_allow_html=True)

thin_divider()

# --- 병합 셀 에러 방지용 커스텀 입력 함수 ---
def safe_write_to_cell(ws, cell_coord, value):
    """
    지정된 좌표(예: 'D9')가 병합된 셀에 속해있을 경우, 
    자동으로 병합 범위의 대표 셀(Top-Left)을 찾아 값을 입력합니다.
    """
    cell = ws[cell_coord]
    # 'MergedCell' 타입이면 껍데기 셀이라는 뜻입니다.
    if type(cell).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                # 병합 범위의 진짜 대표 셀(최소 행/열)에 값 입력
                master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                master_cell.value = value
                return
    else:
        # 일반 셀이면 그냥 입력합니다.
        cell.value = value

def safe_write_rc(ws, r, c, value):
    """ 행(row)과 열(col) 숫자로 셀에 안전하게 값을 입력합니다. """
    coord = ws.cell(row=r, column=c).coordinate
    safe_write_to_cell(ws, coord, value)

# --- 깃허브 저장소의 엑셀 템플릿 매핑 함수 ---
def generate_excel():
    template_path = "지출결의서_양식.xlsx" 
    
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "서버(깃허브)에 '지출결의서_양식.xlsx' 파일이 없습니다."
    else:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # [수정 완료] 이제 껍데기 셀을 지정하더라도 알아서 진짜 셀을 찾아 입력합니다.
        # 첨부해주신 데이터를 토대로 대략적인 좌표를 E8, E9 등으로 조정해두었습니다.
        safe_write_to_cell(ws, 'B6', project)       
        safe_write_to_cell(ws, 'B7', purpose)       
        safe_write_to_cell(ws, 'B8', department)    
        safe_write_to_cell(ws, 'E8', title)         
        safe_write_to_cell(ws, 'G8', account)       
        safe_write_to_cell(ws, 'B9', author)        
        safe_write_to_cell(ws, 'E9', date.strftime('%Y-%m-%d'))
        
        # 지출 내역 반복 입력
        start_row = 13
        for i, row in edited_df.iterrows():
            current_row = start_row + i
            safe_write_rc(ws, current_row, 1, row['지출일'].strftime('%Y-%m-%d')) # A열
            safe_write_rc(ws, current_row, 2, row['적요'])                        # B열
            safe_write_rc(ws, current_row, 4, row['지급처'])                      # D열 (C, D 병합일 수 있음)
            safe_write_rc(ws, current_row, 6, row['금액'])                        # F열
            safe_write_rc(ws, current_row, 8, row['비고'])                        # H열

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

st.subheader("📥 문서 출력")

if st.button("🔄 최종 지출결의서 엑셀 변환", type="primary", use_container_width=True):
    if project == "선택" or not author:
        st.error("프로젝트와 작성자를 입력해주세요.")
    else:
        excel_data = generate_excel()
        file_name = f"지출결의서_{author if author else '미상'}_{datetime.today().strftime('%Y%m%d')}.xlsx"
        
        st.success("문서 변환이 완료되었습니다. 아래 버튼을 눌러 양식이 적용된 엑셀을 다운로드하세요.")
        st.download_button(
            label="📊 완성된 지출결의서 다운로드 (Excel)",
            data=excel_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
